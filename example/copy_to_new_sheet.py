from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import messagebox,filedialog

def main():
    print('excel_read')

    adj_list = [[0 for i in range(4)] for j in range(10)] # 4x10セット

    # filename = filedialog.askopenfilename()
    filename = filedialog.askopenfilename(
        title="slect release test file",
        filetypes=[("excel files", "*.xlsx")],
    )
    if filename == "":
        print('ファイル指定に失敗しました')
        return

    # wb = load_workbook('release_check.xlsx')
    # wb = load_workbook('release_check.xlsx', data_only=True)
    wb = load_workbook(filename, data_only=True)
    # ワークシートの列挙
    for sheet in wb:
        print(f'sheet name: {sheet.title}')

    #ファイル内の全てのシートをループして検索
    target_name = '調整値'
    check = False
    for ws in wb.worksheets:
        #指定シートが存在していれば、変数にTrueを格納
        if ws.title == target_name:
            check = True
    if check == True:
        print(target_name + 'は存在します')
    else:
        print(target_name + 'は存在しません')

    if check == True:
        print('調整値->')
        ws = wb['調整値']  # ワークシートを指定
        # ws = wb.active  # アクティブなワークシートを選択
        i = 0
        for n in range(3,13): # 列(C-L)
            j = 0
            for k in range(37,40): # 行(37-39)
                print(ws.cell(row=k, column=n).value)
                adj_list[i][j] = ws.cell(row=k, column=n).value
                j += 1
            i += 1

    #ファイル内の全てのシートをループして検索
    target_name = 'CRC'
    check = False
    for ws in wb.worksheets:
        #指定シートが存在していれば、変数にTrueを格納
        if ws.title == target_name:
            check = True
    if check == True:
        print(target_name + 'は存在します')
    else:
        print(target_name + 'は存在しません')

    if check == True:
        print('CRC->')
        ws = wb['CRC']  # ワークシートを指定
        for n in range(4,5): # 列(D)
            for k in range(4,14): # 行(4-13)
                print(ws.cell(row=k, column=n).value)

    target_name = 'test'
    check = False
    for ws in wb.worksheets:
        #指定シートが存在していれば、変数にTrueを格納
        if ws.title == target_name:
            check = True
    if check == True:
        print(target_name + 'は存在します')
    else:
        print(target_name + 'は存在しません')
    if check == True:
        # wb.remove_sheet(wb.get_sheet_by_name(target_name))
        # wb.remove_sheet(wb[target_name])
        wb.remove(wb[target_name])
        print(target_name + 'を削除しました')

    # テスト項目のシートを作成する
    wb.create_sheet(title='test')
    ws = wb['test']  # ワークシートを指定
    # セルに書き込み
    ws['A1'] = 'TEMP DATA'
    # ws.cell(row=37, column=3).value = 'dummy' #C37
    # ws.cell(row=37, column=3).value = adj_list[0][0] #C37
    # check = False
    # if check == True:
    #     i = 0
    #     for n in range(3,13): # 列(C-L)
    #         j = 0
    #         for k in range(37,40): # 行(37-39)
    #             ws.cell(row=k, column=n).value = adj_list[i][j] # test
    #             # ws.cell(row=k, column=n).value = str(adj_list[i][j]) # test
    #             print('copy: [' + str(i) + '][' + str(j) + '] ->(' + str(n) + ',' + str(k) + ') :' + str(adj_list[i][j]))
    #             # print('copy: ' + str(adj_list[i][j]))
    #             j += 1
    #         i += 1
    i = 0
    for n in range(3,13): # 列(C-L)
        j = 0
        for k in range(37,40): # 行(37-39)
            ws.cell(row=k, column=n).value = adj_list[i][j] # test
            # ws.cell(row=k, column=n).value = str(adj_list[i][j]) # test
            print('copy: [' + str(i) + '][' + str(j) + '] ->(' + str(n) + ',' + str(k) + ') :' + str(adj_list[i][j]))
            # print('copy: ' + str(adj_list[i][j]))
            j += 1
        i += 1
    try:
        wb.save(filename)  # overwrite excel file
        # check = True
    except:
        print(filename + 'の上書きに失敗しました')

if __name__ == '__main__':
    main()