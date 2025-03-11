from openpyxl import Workbook
from openpyxl import load_workbook

def main():
    print('excel_read')
    # wb = load_workbook('release_check.xlsx')
    wb = load_workbook('release_check.xlsx', data_only=True)
    # ワークシートの列挙
    for sheet in wb:
        print(f'sheet name: {sheet.title}')

    print('調整値->')
    ws = wb['調整値']  # ワークシートを指定
    # ws = wb.active  # アクティブなワークシートを選択
    for n in range(3,13): # 列(C-L)
        for i in range(37,40): # 行(37-39)
            print(ws.cell(row=i, column=n).value)

    print('CRC->')
    ws = wb['CRC']  # ワークシートを指定
    for n in range(4,5): # 列(D)
        for i in range(4,14): # 行(4-13)
            print(ws.cell(row=i, column=n).value)

if __name__ == '__main__':
    main()