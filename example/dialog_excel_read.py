from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import messagebox,filedialog

def main():
    print('excel_read')

    # filename = filedialog.askopenfilename()
    filename = filedialog.askopenfilename(
        title="slect release test file",
        filetypes=[("excel files", "*.xlsx")],
    )
    if filename == "":
        print('cancel')
        return

    # wb = load_workbook('release_check.xlsx')
    # wb = load_workbook('release_check.xlsx', data_only=True)
    wb = load_workbook(filename, data_only=True)
    # ワークシートの列挙
    for sheet in wb:
        print(f'sheet name: {sheet.title}')

    #ファイル内の全てのシートをループして検索
    target_name = '調整値'
    check = False
    for ws in wb.worksheets:
        #指定シートが存在していれば、変数にTrueを格納
        if ws.title == target_name:
            check = True
    if check == True:
        print(target_name + 'は存在します')
    else:
        print(target_name + 'は存在しません')

    if check == True:
        print('調整値->')
        ws = wb['調整値']  # ワークシートを指定
        # ws = wb.active  # アクティブなワークシートを選択
        for n in range(3,13): # 列(C-L)
            for i in range(37,40): # 行(37-39)
                print(ws.cell(row=i, column=n).value)

    #ファイル内の全てのシートをループして検索
    target_name = 'CRC'
    check = False
    for ws in wb.worksheets:
        #指定シートが存在していれば、変数にTrueを格納
        if ws.title == target_name:
            check = True
    if check == True:
        print(target_name + 'は存在します')
    else:
        print(target_name + 'は存在しません')

    if check == True:
        print('CRC->')
        ws = wb['CRC']  # ワークシートを指定
        for n in range(4,5): # 列(D)
            for i in range(4,14): # 行(4-13)
                print(ws.cell(row=i, column=n).value)

if __name__ == '__main__':
    main()